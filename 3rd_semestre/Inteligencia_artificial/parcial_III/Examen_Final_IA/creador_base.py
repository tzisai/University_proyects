import time
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.keys import Keys
from io import StringIO
from selenium.webdriver.firefox.options import Options
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime

# Ruta del ejecutable de Firefox
firefox_binary_path = r"C:\Program Files\Mozilla Firefox\firefox.exe"  # Cambia esta ruta según tu instalación

# Configura las opciones de Firefox
options = Options()
options.binary_location = firefox_binary_path  # Establece la ubicación de Firefox
options.add_argument("--start-maximized")  # Maximiza la ventana
options.add_argument("--disable-extensions")  # Deshabilita extensiones

# Configura el servicio de geckodriver (esto se encargará de descargar y usar la versión adecuada)
service = FirefoxService(executable_path=GeckoDriverManager().install())

def parse_body_content(body_content):
    """
    Parses the body content with the specific format and returns a DataFrame
    """
    try:
        # Remove <body> and </body> tags and any script tags
        content = re.sub(r'</?body>|<script.*?</script>', '', body_content, flags=re.DOTALL).strip()

        # Split the content by <br> tags and clean up
        rows = [row.strip() for row in content.split('<br>') if row.strip()]

        # Process each row to create a list of dictionaries
        data = []
        for row in rows:
            # Skip rows that don't look like data
            if not re.match(r'^[A-Za-z]+,\s*\d+.*', row):
                continue

            values = [v.strip() for v in row.split(',')]
            if len(values) >= 3:  # Asegurarse de que hay al menos 3 valores
                try:
                    data.append({
                        'Amino_Acid': values[0],
                        'Count': int(values[1]),
                        'Percentage': float(values[2])
                    })
                except (ValueError, IndexError):
                    continue  # Skip rows with invalid data

        if not data:  # Si no se encontraron datos válidos
            print("No se encontraron datos válidos en el contenido")
            return None

        # Create DataFrame
        return pd.DataFrame(data)

    except Exception as e:
        print(f"Error parsing body content: {e}")
        return None

def hacer_scraping(driver, elemento):
    """
    Realiza web scraping para un elemento específico
    Devuelve un DataFrame con los resultados o None si falla
    """
    try:
        # Ir a la página de búsqueda de NCBI con el elemento ingresado
        search_url = f"https://www.ncbi.nlm.nih.gov/search/all/?term={elemento}"
        driver.get(search_url)
        print(f"Navegando a la búsqueda de {elemento}...")

        # Espera explícita para el enlace de Nucleótidos
        wait = WebDriverWait(driver, 20)  # Aumentado a 20 segundos
        try:
            nucleotide_link = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//span[@class='nwds-list-item__primary-text' and text()='Nucleotide']"))
            )
            nucleotide_link.click()
            print("Enlace Nucleotide encontrado y clickeado")
        except Exception as e:
            print(f"Error al encontrar/hacer clic en enlace Nucleotide: {e}")
            return None

        # Espera a que cargue la página de Nucleótidos
        time.sleep(5)  # Aumentado a 5 segundos

        # Buscar el primer texto que coincida con /translation="
        page_source = driver.page_source
        translation_match = re.search(r'/translation="([^"]+)"', page_source)

        if not translation_match:
            print(f"No se encontró traducción para {elemento}")
            return None

        translation_text = translation_match.group(1)
        print(f"Traducción para {elemento} obtenida exitosamente.")

        # Navegar a la página de ExPASy ProtParam
        print("Navegando a ExPASy ProtParam...")
        driver.get("https://web.expasy.org/protparam/")
        time.sleep(5)  # Aumentado a 5 segundos

        # Encontrar el campo de texto y pegar la secuencia
        try:
            textarea = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, "sequence"))
            )
            textarea.clear()
            textarea.send_keys(translation_text)
            print("Secuencia pegada exitosamente")
        except Exception as e:
            print(f"Error al pegar la secuencia: {e}")
            return None

        # Hacer clic en el botón "Compute parameters"
        try:
            compute_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@value='Compute parameters']"))
            )
            compute_button.click()
            print("Parámetros computados")
            time.sleep(5)
        except Exception as e:
            print(f"Error al hacer clic en Compute parameters: {e}")
            return None

        # Hacer clic en el botón "CSV format"
        try:
            csv_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@value='CSV format']"))
            )
            csv_button.click()
            print("Formato CSV solicitado")
            time.sleep(3)
        except Exception as e:
            print(f"Error al hacer clic en CSV format: {e}")
            return None

        # Obtener contenido del body
        try:
            body_content = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            ).get_attribute('innerHTML')
        except Exception as e:
            print(f"Error al obtener el contenido del body: {e}")
            return None

        # Parsear el contenido del body
        df = parse_body_content(body_content)

        if df is not None:
            df['Elemento'] = elemento  # Añadir columna con el elemento original
            print(f"Datos procesados exitosamente para {elemento}")
            return df
        else:
            print(f"No se pudieron procesar los datos para {elemento}")
            return None

    except Exception as e:
        print(f"Error al procesar {elemento}: {e}")
        return None

# El resto de las funciones (obtener_elementos, guardar_datos, main) permanecen igual
def obtener_elementos(file_path):
    """
    Lee elementos del archivo, excluyendo líneas con '#'
    """
    elementos = []
    try:
        with open(file_path, 'r') as file:
            for line in file:
                elemento = line.split('#')[0].strip()
                if elemento:
                    elementos.append(elemento)
        return elementos
    except FileNotFoundError:
        print(f"Error: El archivo {file_path} no se encontró.")
        return []
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
        return []


def guardar_datos(df, filename):
    """
    Guarda un DataFrame en un archivo Excel (.xlsx) o CSV (.csv), dependiendo de la extensión.
    """
    try:
        if filename.endswith('.xlsx'):
            try:
                # Intentar cargar el archivo existente
                try:
                    workbook = load_workbook(filename)
                except FileNotFoundError:
                    # Si el archivo no existe, crear uno nuevo directamente
                    df.to_excel(filename, sheet_name="Datos", index=False)
                    print(f"Archivo {filename} creado exitosamente.")
                    return
                except Exception as e:
                    # Si hay cualquier otro error al cargar el archivo, crear uno nuevo
                    print(f"Error al cargar el archivo existente: {e}")
                    print("Creando nuevo archivo...")
                    df.to_excel(filename, sheet_name="Datos", index=False)
                    print(f"Archivo {filename} creado exitosamente.")
                    return

                # Si el archivo existe y se cargó correctamente
                sheet_name = "Datos"
                if sheet_name not in workbook.sheetnames:
                    workbook.create_sheet(sheet_name)

                sheet = workbook[sheet_name]
                last_row = sheet.max_row

                # Convertir DataFrame a lista de listas
                data_to_write = [df.columns.tolist()] if last_row == 1 else []
                data_to_write.extend(df.values.tolist())

                # Escribir los datos fila por fila
                for row_data in data_to_write:
                    sheet.append(row_data)

                # Guardar el workbook
                workbook.save(filename)
                print(f"Datos guardados en {filename} exitosamente.")

            except Exception as e:
                print(f"Error al procesar el archivo Excel: {e}")
                # Intentar guardar como nuevo archivo
                try:
                    df.to_excel(filename, sheet_name="Datos", index=False, mode='w')
                    print(f"Datos guardados en nuevo archivo {filename}")
                except Exception as e_new:
                    print(f"Error al crear nuevo archivo: {e_new}")
                    # Intentar con un nombre de archivo alternativo
                    alt_filename = f"backup_{filename}"
                    df.to_excel(alt_filename, sheet_name="Datos", index=False)
                    print(f"Datos guardados en archivo alternativo: {alt_filename}")

        elif filename.endswith('.csv'):
            # Para archivos CSV
            exists = pd.io.common.file_exists(filename)
            df.to_csv(filename, mode='a', index=False, header=not exists)
            print(f"Datos guardados en {filename} exitosamente.")

        else:
            print("Error: Extensión de archivo no soportada. Utilice '.xlsx' o '.csv'.")

    except Exception as e:
        print(f"Error al guardar los datos: {e}")
        # Intentar guardar en un archivo de respaldo
        backup_filename = f"backup_{int(time.time())}_{filename}"
        try:
            df.to_excel(backup_filename, sheet_name="Datos", index=False)
            print(f"Datos guardados en archivo de respaldo: {backup_filename}")
        except Exception as e_backup:
            print(f"Error al guardar archivo de respaldo: {e_backup}")
            # Último recurso: guardar como CSV
            csv_backup = f"backup_{int(time.time())}.csv"
            df.to_csv(csv_backup, index=False)
            print(f"Datos guardados en CSV de respaldo: {csv_backup}")

def main():
    driver = None
    try:
        # Configurar opciones de Firefox
        firefox_binary_path = r"C:\Program Files\Mozilla Firefox\firefox.exe"
        options = Options()
        options.binary_location = firefox_binary_path
        options.add_argument("--start-maximized")
        options.add_argument("--disable-extensions")

        # Configurar el servicio de geckodriver con log silencioso
        service = FirefoxService(
            executable_path=GeckoDriverManager().install(),
            log_output="geckodriver.log"  # Agregar log para diagnóstico
        )

        # Crear el controlador de Firefox con un tiempo de espera explícito
        driver = webdriver.Firefox(service=service, options=options)
        driver.implicitly_wait(10)  # Espera implícita global de 10 segundos

        file_path = r'C:\Users\jaquy\Documents\UAA\3_semestre\Inteligencia artificial\Examen FInal\Database.txt'
        output_file = 'dataset_protparam.xlsx'

        elementos = obtener_elementos(file_path)
        if not elementos:
            print("No se encontraron elementos para procesar")
            return

        print(f"Elementos encontrados: {elementos}")

        for elemento in elementos:
            try:
                print(f"\nProcesando elemento: {elemento}")
                df = hacer_scraping(driver, elemento)
                if df is not None:
                    guardar_datos(df, output_file)
                time.sleep(2)  # Pequeña pausa entre elementos
            except Exception as e:
                print(f"Error procesando elemento {elemento}: {e}")
                continue  # Continuar con el siguiente elemento si hay error

    except Exception as e:
        print(f"Error general en el script: {e}")

    finally:
        # Asegurarse de cerrar el driver apropiadamente
        if driver:
            try:
                driver.quit()
                print("Navegador cerrado correctamente")
            except Exception as e:
                print(f"Error al cerrar el navegador: {e}")


if _name_ == "_main_":
    main()